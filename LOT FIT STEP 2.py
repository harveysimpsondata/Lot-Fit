import os
import openpyxl
from openpyxl.utils import get_column_letter
os.chdir(os.getcwd())
wb = openpyxl.load_workbook('1_Lot Fit Matrix with Exposed Columns.xlsx')
sheet=wb['Lot Fit Matrix']
sheet2=wb['Length']





House_List = ['Tea Olive', 'Oakmonte','Greenfield','Biltmore','Harding','Reynolds']
Side = False




House_Column_List = [] 
all_column_list = []
Side_Column_List=[]

count=1   
count_2=1

#get column values in House_List     
for i in range(3,sheet.max_column):
    
    count+=1
    
    
    for j in House_List:
        
        if j in sheet.cell(row=7,column=count).value:
            
            a=get_column_letter(count)
            
            House_Column_List.append(a)
            
        else:
            pass

#get all column values            
for i in range(3,sheet.max_column):
    
    count+=1
    a=get_column_letter(i)
    all_column_list.append(a)
    
    
    
#get side column values

for i in range(3,sheet.max_column):
    count_2+=1
    if ("Side") in sheet.cell(row=7,column=count_2).value:
        a=get_column_letter(count_2)
        Side_Column_List.append(a)
        
    elif ("Bolt") in sheet.cell(row=7,column=count_2).value:
        a=get_column_letter(count_2)
        Side_Column_List.append(a)
        
    else:
        pass 




if Side == True:
    
    not_column_list = list(set(all_column_list) - set(House_Column_List))

    for col in not_column_list:
        sheet.column_dimensions[col].hidden = True

elif Side == False:
    
    
    house_side_list = []
    
    
    
    for i in Side_Column_List:
        if i in House_Column_List:
            house_side_list.append(i)
        
    house_without_side_list = []

    for i in House_Column_List:
        if i not in house_side_list:
            house_without_side_list.append(i)
    
    not_column_list = list(set(all_column_list) - set(house_without_side_list))
    
 
    for col in not_column_list:
        sheet.column_dimensions[col].hidden = True    
    
wb.save('2_Lot Fit Matrix Final.xlsx')


     