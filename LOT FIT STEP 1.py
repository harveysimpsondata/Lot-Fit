import os
import openpyxl

from openpyxl.styles import PatternFill, Border, Side

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

greenFill = PatternFill(start_color='00FF00',
                   end_color='00FF00',
                   fill_type='solid')

thin_border = Border(top=Side(style='thin'), left=Side(style='thin'),
                     right=Side(style='thin'), bottom=Side(style='thin'))

os.chdir(os.getcwd())
wb = openpyxl.load_workbook('Lot Fit Matrix.xlsx')
sheet=wb['Lot Fit Matrix']
sheet2=wb['Length']

class House:
    
    def __init__(self,width,length,side,coveredPorch):
        self.width=width;
        self.length=length;
        self.side=side
        self.cp=coveredPorch
    pass



#project
Grey_Birch = House(39,54,0,64)
Silver_Maple = House(39,70,0,70)
Sycamore = House(39,51,0,57)
Jasmine = House(40,62,0,74)
Yellow_Jasmine = House(40,70,0,70)
Mitchell = House(29,47,0,57)
Kephart = House(29,44,0,54)
Barkley = House(39,54,0,64)


Allatoona_A2_C2_C3 = House(39,63,0,73)
Allatoona_B3 = House(39,69.06,0,79.06)

#Ambrose
Ambrose_A = House(39,56,0,66)
Ambrose_B=House(39,62,0,66)
Ambrose_C=House(39,63.5,0,73.5)


#Barrington
Barrington_A2_D2_E2_F2 = House(50,57,0,57)
Barrington_B2_C2 = House(50,56,0,56)
Barrington_A2_D2_E2_F2_Side = House(50,57,28,57)
Barrington_B2_C2_Side = House(50,56,28,56)
Barrington_A2_D2_E2_F2_BoltOn = House(62,57,16,57)
Barrington_B2_C2_BoltOn = House(62,56,16,56)

#Biltmore
Biltmore_A2_B2_C2_D2 = House(50,43,28,55)
Biltmore_A2_B2_C2_D2_BoltOn = House(62,43,16,55)


#Burton
Burton_C3_F3_T3 = House(37,42,0,52)

#Centennial
Centennial_A_B_C = House(34,38.5,0,48.5)

#Crestwood
Crestwood_A_B_C_D_E_F = House(39,66,0,66)

#cumberland
Cumberland_A_B_C = House(38,62,0,74)

#Dakota
Dakota_A_B_E = House(40,49.67,28,61.67)
Dakota_C_D = House(40,49.17,28,61.17)

Dakota_A_B_E_3_Front = House(50,49.67,0,61.67)
Dakota_C_D_3_Front = House(50,49.17,0,61.17)

#Firethorn
Firethorn_A_B_C_M_N_O = House(49,65,28,65)

#Fontana
Fontana_C3_F3_T3 = House(39,41.33,0,51.33)
Fontana_C3_F3_T3_ext = House(39,41.33,0,53.33)

#Greenfield
Greenfield_C3_F3_T3 = House(30,46.25,0,56.25)
Greenfield_C3_F3_T3_ext = House(30,46.25,0,56.25)

#Harding
Harding_C3_F3_T3 = House(39,49.5,0,59.5)
Harding_C3_F3_T3_ext = House(39,49.5,0,61.5)

#Jordan
Jordan_A1_B2_C3 = House(39,50.5,0,60.5)

#Juniper
Juniper_A_B_D_E=House(40,54,28,66)
Juniper_C = House(40,56,28,68)
Juniper_A_B_D_E_3_Front = House(50,54,0,66)
Juniper_C_3_Front=House(50,56,0,68)

#Mcdowell
Mcdowell_A2 = House(34,62,0,72)
Mcdowell_B2_C2 = House(34,64,0,74)

#Middleton
Middleton_A2_D2 = House(40,60,28,70)
Middleton_B2_C2 = House(40,59,28,69)

#Nandina
Nandina_A_M = House(49,72,0,73.5)
Nandina_B_C_N_O = House(49,74.17,0,78.17)
Nandina_A_M_3_Front = House(61,72,0,76)
Nandina_B_C_N_O_3_Front = House(61,74.17,0,78.17)
Nandina_A_M_Side = House(49,72,28,73.5)
Nandina_B_C_N_O_Side = House(49,74.17,28,78.17)
Nandina_A_M_3_BoltOn = House(61,72,16,76)
Nandina_B_C_N_O_3_BoltOn = House(61,74.17,16,78.17)

#Oakmonte
Oakmonte_A = House(50,51,0,57)
Oakmonte_B_C_D_E_F_G_H = House(50,50,0,56)
Oakmonte_A_Side = House(50,51,28,57)
Oakmonte_B_C_D_E_F_G_H_Side = House(50,50,28,56)
Oakmonte_A_BoltOn = House(62,51,16,57)
Oakmonte_B_C_D_E_F_G_H_BoltOn = House(62,50,16,56)

#Reynolds
Reynolds_A2 = House(50,49,0,59)
Reynolds_B2 = House(50,48,0,58)
Reynolds_C2 = House(50,48.5,0,58.5)
Reynolds_D2 = House(50,47,0,57)
Reynolds_A2_Side = House(50,49,28,59)
Reynolds_B2_Side = House(50,48,28,58)
Reynolds_C2_Side = House(50,48.5,28,58.5)
Reynolds_D2_Side = House(50,47,28,57)
Reynolds_A2_BoltOn = House(62,49,16,59)
Reynolds_B2_BoltOn = House(62,48,16,58)
Reynolds_C2_BoltOn = House(62,48.5,16,58.5)
Reynolds_D2_BoltOn = House(62,47,16,57)

#Sienna
Sienna_A = House(39,74,0,86)
Sienna_B_C = House(39,76,0,88)

#Sinclair
Sinclair_A1_A3_B1_B3_C1_C3 = House(39,59.25,0,69.25)

#Tea Olive
Tea_Olive_A_B = House(49,67,0,67)
Tea_Olive_C = House(49,71.42,0,71.42)
Tea_Olive_A_B_Side = House(49,67,28,67)
Tea_Olive_C_Side = House(49,71.42,28,71.42)
Tea_Olive_A_B_BoltOn = House(61,67,16,67)
Tea_Olive_C_BoltOn = House(61,71.42,16,71.42)

#Willow
Willow_A2 = House(50,48,28,60)
Willow_B2_C2_D2 = House(50,50.33,28,62.33)
Willow_E2 = House(50,50,28,62)
Willow_F2 = House(50,48.5,28,60.5)
Willow_G2 = House(50,49,28,61)

#Willow_A2_Side = House(50,48,28,60)
#Willow_B2_C2_D2_Side = House(50,50.33,28,62.33)
#Willow_E2_Side = House(50,50,28,62)
#Willow_F2_Side = House(50,48.5,28,60.5)
#Willow_G2_Side = House(50,49,28,61)

Willow_A2_BoltOn = House(62,48,16,60)
Willow_B2_C2_D2_BoltOn = House(62,50.33,16,62.33)
Willow_E2_BoltOn = House(62,50,16,62)
Willow_F2_BoltOn = House(62,48.5,16,60.5)
Willow_G2_BoltOn = House(62,49,16,61)

#Winston
Winston_A_B_C = House(40,46,28,58)
Winston_A_B_C_3_Front=House(50,46,0,58)




count=1        
for i in range(2,sheet.max_column):
    
    count+=1
               
    if sheet.cell(row=7,column=count).value == "Allatoona A & C w/ Patio":
        sheet.cell(row=3,column=count).value = Allatoona_A2_C2_C3.width
        sheet.cell(row=4,column=count).value = Allatoona_A2_C2_C3.length
        for j in range(8, sheet.max_row+1):
                           
            if Allatoona_A2_C2_C3.width <= sheet2.cell(row=j,column=2).value and Allatoona_A2_C2_C3.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Allatoona_A2_C2_C3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Allatoona A & C w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")

    elif sheet.cell(row=7,column=count).value == "Allatoona A & C w/ Covered Patio":
        sheet.cell(row=3,column=count).value = Allatoona_A2_C2_C3.width
        sheet.cell(row=4,column=count).value = Allatoona_A2_C2_C3.cp
        for j in range(8, sheet.max_row+1):
                           
            if Allatoona_A2_C2_C3.width <= sheet2.cell(row=j,column=2).value and Allatoona_A2_C2_C3.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Allatoona_A2_C2_C3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Allatoona A & C w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Allatoona B w/ Patio":
        sheet.cell(row=3,column=count).value = Allatoona_B3.width
        sheet.cell(row=4,column=count).value = Allatoona_B3.length
        for j in range(8, sheet.max_row+1):
                           
            if Allatoona_B3.width <= sheet2.cell(row=j,column=2).value and Allatoona_B3.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Allatoona_B3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Allatoona B w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Allatoona B w/ Covered Patio":
        sheet.cell(row=3,column=count).value = Allatoona_B3.width
        sheet.cell(row=4,column=count).value = Allatoona_B3.cp
        for j in range(8, sheet.max_row+1):
                           
            if Allatoona_B3.width <= sheet2.cell(row=j,column=2).value and Allatoona_B3.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Allatoona_B3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Allatoona B w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Grey Birch":
        sheet.cell(row=3,column=count).value = Grey_Birch.width
        sheet.cell(row=4,column=count).value = Grey_Birch.length
        for j in range(8, sheet.max_row+1):
                           
            if Grey_Birch.width <= sheet2.cell(row=j,column=2).value and Grey_Birch.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Grey_Birch.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Grey Birch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Silver Maple":
        sheet.cell(row=3,column=count).value = Silver_Maple.width
        sheet.cell(row=4,column=count).value = Silver_Maple.length
        for j in range(8, sheet.max_row+1):
                           
            if Silver_Maple.width <= sheet2.cell(row=j,column=2).value and Silver_Maple.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Silver_Maple.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Silver Maple does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Sycamore":
        sheet.cell(row=3,column=count).value = Sycamore.width
        sheet.cell(row=4,column=count).value = Sycamore.length
        for j in range(8, sheet.max_row+1):
                           
            if Sycamore.width <= sheet2.cell(row=j,column=2).value and Sycamore.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Sycamore.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Sycamore does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Jasmine":
        sheet.cell(row=3,column=count).value = Jasmine.width
        sheet.cell(row=4,column=count).value = Jasmine.length
        for j in range(8, sheet.max_row+1):
                           
            if Jasmine.width <= sheet2.cell(row=j,column=2).value and Jasmine.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Jasmine.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Jasmine does not fit in lot {sheet.cell(row=j, column=1).value}")
            
    elif sheet.cell(row=7,column=count).value == "Yellow Jasmine":
        sheet.cell(row=3,column=count).value = Yellow_Jasmine.width
        sheet.cell(row=4,column=count).value = Yellow_Jasmine.length
        for j in range(8, sheet.max_row+1):
                           
            if Yellow_Jasmine.width <= sheet2.cell(row=j,column=2).value and Yellow_Jasmine.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Yellow_Jasmine.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Yellow Jasmine does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Mitchell":
        sheet.cell(row=3,column=count).value = Mitchell.width
        sheet.cell(row=4,column=count).value = Mitchell.length
        for j in range(8, sheet.max_row+1):
                           
            if Mitchell.width <= sheet2.cell(row=j,column=2).value and Mitchell.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Mitchell.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Mitchell does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Kephart":
        sheet.cell(row=3,column=count).value = Kephart.width
        sheet.cell(row=4,column=count).value = Kephart.length
        for j in range(8, sheet.max_row+1):
                           
            if Kephart.width <= sheet2.cell(row=j,column=2).value and Kephart.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Kephart.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Kephart does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Barkley":
        sheet.cell(row=3,column=count).value = Barkley.width
        sheet.cell(row=4,column=count).value = Barkley.length
        for j in range(8, sheet.max_row+1):
                           
            if Barkley.width <= sheet2.cell(row=j,column=2).value and Barkley.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barkley.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barkley does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Ambrose A w/ Patio":
        sheet.cell(row=3,column=count).value = Ambrose_A.width
        sheet.cell(row=4,column=count).value = Ambrose_A.length
        for j in range(8, sheet.max_row+1):
                           
            if Ambrose_A.width <= sheet2.cell(row=j,column=2).value and Ambrose_A.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Ambrose_A.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Ambrose A w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Ambrose A w/ Screened Covered Porch") or (sheet.cell(row=7,column=count).value == "Ambrose A w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Ambrose A w/ Morning Room Screen/Covered Patio") or (sheet.cell(row=7,column=count).value == "Ambrose A w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Ambrose_A.width
        sheet.cell(row=4,column=count).value = Ambrose_A.cp
        for j in range(8, sheet.max_row+1):
                           
            if Ambrose_A.width <= sheet2.cell(row=j,column=2).value and Ambrose_A.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Ambrose_A.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Ambrose A w/ Screened Covered Porch, Ambrose A w/ Morning Room, Ambrose A w/ Morning Room Screen/Covered Patio, and Ambrose A w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Ambrose B w/ Patio":
        sheet.cell(row=3,column=count).value = Ambrose_B.width
        sheet.cell(row=4,column=count).value = Ambrose_B.length
        for j in range(8, sheet.max_row+1):
                           
            if Ambrose_B.width <= sheet2.cell(row=j,column=2).value and Ambrose_B.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Ambrose_B.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Ambrose B w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Ambrose B w/ Screened Covered Porch") or (sheet.cell(row=7,column=count).value == "Ambrose B w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Ambrose B w/ Morning Room Screen/Covered Patio") or (sheet.cell(row=7,column=count).value == "Ambrose B w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Ambrose_B.width
        sheet.cell(row=4,column=count).value = Ambrose_B.cp
        for j in range(8, sheet.max_row+1):
                           
            if Ambrose_B.width <= sheet2.cell(row=j,column=2).value and Ambrose_B.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Ambrose_B.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Ambrose B w/ Screened Covered Porch, Ambrose B w/ Morning Room, Ambrose B w/ Morning Room Screen/Covered Patio, and Ambrose B w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif sheet.cell(row=7,column=count).value == "Ambrose C w/ Patio":
        sheet.cell(row=3,column=count).value = Ambrose_C.width
        sheet.cell(row=4,column=count).value = Ambrose_C.length
        for j in range(8, sheet.max_row+1):
                           
            if Ambrose_C.width <= sheet2.cell(row=j,column=2).value and Ambrose_C.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Ambrose_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Ambrose C w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Ambrose C w/ Screened Covered Porch") or (sheet.cell(row=7,column=count).value == "Ambrose C w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Ambrose C w/ Morning Room Screen/Covered Patio") or (sheet.cell(row=7,column=count).value == "Ambrose C w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Ambrose_C.width
        sheet.cell(row=4,column=count).value = Ambrose_C.cp
        for j in range(8, sheet.max_row+1):
                           
            if Ambrose_C.width <= sheet2.cell(row=j,column=2).value and Ambrose_C.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Ambrose_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Ambrose C w/ Screened Covered Porch, Ambrose C w/ Morning Room, Ambrose C w/ Morning Room Screen/Covered Patio, and Ambrose C w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington A, D, E, & F w/ Patio"):
        sheet.cell(row=3,column=count).value = Barrington_A2_D2_E2_F2.width
        sheet.cell(row=4,column=count).value = Barrington_A2_D2_E2_F2.length
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_A2_D2_E2_F2.width <= sheet2.cell(row=j,column=2).value and Barrington_A2_D2_E2_F2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_A2_D2_E2_F2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington A, D, E, & F w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington A, D, E, & F w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Barrington_A2_D2_E2_F2.width
        sheet.cell(row=4,column=count).value = Barrington_A2_D2_E2_F2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_A2_D2_E2_F2.width <= sheet2.cell(row=j,column=2).value and Barrington_A2_D2_E2_F2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_A2_D2_E2_F2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington A, D, E, & F w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington B & C w/ Patio"):
        sheet.cell(row=3,column=count).value = Barrington_B2_C2.width
        sheet.cell(row=4,column=count).value = Barrington_B2_C2.length
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_B2_C2.width <= sheet2.cell(row=j,column=2).value and Barrington_B2_C2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_B2_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington B & C w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington B & C w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Barrington_B2_C2.width
        sheet.cell(row=4,column=count).value = Barrington_B2_C2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_B2_C2.width <= sheet2.cell(row=j,column=2).value and Barrington_B2_C2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_B2_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington B & C w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington A, D, E, & F Side Entry w/ Patio"):
        sheet.cell(row=3,column=count).value = Barrington_A2_D2_E2_F2_Side.width
        sheet.cell(row=4,column=count).value = Barrington_A2_D2_E2_F2_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_A2_D2_E2_F2_Side.width <= sheet2.cell(row=j,column=2).value and Barrington_A2_D2_E2_F2_Side.length <= sheet2.cell(row=j,column=3).value and Barrington_A2_D2_E2_F2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Barrington_A2_D2_E2_F2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_A2_D2_E2_F2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington A, D, E, & F Side Entry w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington A, D, E, & F Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Barrington_A2_D2_E2_F2_Side.width
        sheet.cell(row=4,column=count).value = Barrington_A2_D2_E2_F2_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_A2_D2_E2_F2_Side.width <= sheet2.cell(row=j,column=2).value and Barrington_A2_D2_E2_F2_Side.cp <= sheet2.cell(row=j,column=3).value and Barrington_A2_D2_E2_F2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Barrington_A2_D2_E2_F2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_A2_D2_E2_F2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington A, D, E, & F Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington B & C Side Entry w/ Patio"):
        sheet.cell(row=3,column=count).value = Barrington_B2_C2_Side.width
        sheet.cell(row=4,column=count).value = Barrington_B2_C2_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_B2_C2_Side.width <= sheet2.cell(row=j,column=2).value and Barrington_B2_C2_Side.length <= sheet2.cell(row=j,column=3).value and Barrington_B2_C2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Barrington_B2_C2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_B2_C2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington B & C Side Entry w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington B & C Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Barrington_B2_C2_Side.width
        sheet.cell(row=4,column=count).value = Barrington_B2_C2_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_B2_C2_Side.width <= sheet2.cell(row=j,column=2).value and Barrington_B2_C2_Side.cp <= sheet2.cell(row=j,column=3).value and Barrington_B2_C2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Barrington_B2_C2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_B2_C2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington B & C Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington A, D, E, & F Bolt On w/ Patio"):
        sheet.cell(row=3,column=count).value = Barrington_A2_D2_E2_F2_BoltOn.width
        sheet.cell(row=4,column=count).value = Barrington_A2_D2_E2_F2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_A2_D2_E2_F2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Barrington_A2_D2_E2_F2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Barrington_A2_D2_E2_F2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Barrington_A2_D2_E2_F2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_A2_D2_E2_F2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington A, D, E, & F Bolt On w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington A, D, E, & F Bolt On w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Barrington_A2_D2_E2_F2_BoltOn.width
        sheet.cell(row=4,column=count).value = Barrington_A2_D2_E2_F2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_A2_D2_E2_F2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Barrington_A2_D2_E2_F2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Barrington_A2_D2_E2_F2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Barrington_A2_D2_E2_F2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_A2_D2_E2_F2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington A, D, E, & F Bolt On w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington B & C Bolt On w/ Patio"):
        sheet.cell(row=3,column=count).value = Barrington_B2_C2_BoltOn.width
        sheet.cell(row=4,column=count).value = Barrington_B2_C2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_B2_C2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Barrington_B2_C2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Barrington_B2_C2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Barrington_B2_C2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_B2_C2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington B & C Bolt On w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Barrington B & C Bolt On w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Barrington_B2_C2_BoltOn.width
        sheet.cell(row=4,column=count).value = Barrington_B2_C2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Barrington_B2_C2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Barrington_B2_C2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Barrington_B2_C2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Barrington_B2_C2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Barrington_B2_C2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Barrington B & C Bolt On w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D w/ Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Biltmore_A2_B2_C2_D2.width
        sheet.cell(row=4,column=count).value = Biltmore_A2_B2_C2_D2.length
        for j in range(8, sheet.max_row+1):
                           
            if Biltmore_A2_B2_C2_D2.width <= sheet2.cell(row=j,column=2).value and Biltmore_A2_B2_C2_D2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Biltmore_A2_B2_C2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Biltmore A, B, C, & D w/ Patio and Biltmore A, B, C, & D w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D w/ Morning Room Porch"):
        sheet.cell(row=3,column=count).value = Biltmore_A2_B2_C2_D2.width
        sheet.cell(row=4,column=count).value = Biltmore_A2_B2_C2_D2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Biltmore_A2_B2_C2_D2.width <= sheet2.cell(row=j,column=2).value and Biltmore_A2_B2_C2_D2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Biltmore_A2_B2_C2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Biltmore A, B, C, & D w/ Covered Patio, Biltmore A, B, C, & D w/ Ext. Covered Patio, Biltmore A, B, C, & D w/ Morning Room, and Biltmore A, B, C, & D w/ Morning Room Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Biltmore_A2_B2_C2_D2.width
        sheet.cell(row=4,column=count).value = Biltmore_A2_B2_C2_D2.length
        for j in range(8, sheet.max_row+1):
                           
            if Biltmore_A2_B2_C2_D2.width <= sheet2.cell(row=j,column=2).value and Biltmore_A2_B2_C2_D2.length <= sheet2.cell(row=j,column=3).value and Biltmore_A2_B2_C2_D2.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Biltmore_A2_B2_C2_D2.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Biltmore_A2_B2_C2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Biltmore A, B, C, & D Side Entry w/ Patio and Biltmore A, B, C, & D Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Side Entry w/ Morning Room Porch"):
        sheet.cell(row=3,column=count).value = Biltmore_A2_B2_C2_D2.width
        sheet.cell(row=4,column=count).value = Biltmore_A2_B2_C2_D2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Biltmore_A2_B2_C2_D2.width <= sheet2.cell(row=j,column=2).value and Biltmore_A2_B2_C2_D2.cp <= sheet2.cell(row=j,column=3).value and Biltmore_A2_B2_C2_D2.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Biltmore_A2_B2_C2_D2.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Biltmore_A2_B2_C2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Biltmore A, B, C, & D Side Entry w/ Covered Patio, Biltmore A, B, C, & D Side Entry w/ Ext. Covered Patio, Biltmore A, B, C, & D Side Entry w/ Morning Room, and Biltmore A, B, C, & D Side Entry w/ Morning Room Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Biltmore_A2_B2_C2_D2_BoltOn.width
        sheet.cell(row=4,column=count).value = Biltmore_A2_B2_C2_D2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Biltmore_A2_B2_C2_D2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Biltmore_A2_B2_C2_D2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Biltmore_A2_B2_C2_D2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Biltmore_A2_B2_C2_D2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Biltmore_A2_B2_C2_D2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Biltmore A, B, C, & D Bolt On w/ Patio and Biltmore A, B, C, & D Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Biltmore A, B, C, & D Bolt On w/ Morning Room Porch"):
        sheet.cell(row=3,column=count).value = Biltmore_A2_B2_C2_D2_BoltOn.width
        sheet.cell(row=4,column=count).value = Biltmore_A2_B2_C2_D2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Biltmore_A2_B2_C2_D2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Biltmore_A2_B2_C2_D2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Biltmore_A2_B2_C2_D2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Biltmore_A2_B2_C2_D2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Biltmore_A2_B2_C2_D2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Biltmore A, B, C, & D Bolt On w/ Covered Patio, Biltmore A, B, C, & D Bolt On w/ Ext. Covered Patio, Biltmore A, B, C, & D Bolt On w/ Morning Room, and Biltmore A, B, C, & D Bolt On w/ Morning Room Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Burton C, F, & T w/ Patio"):
        sheet.cell(row=3,column=count).value = Burton_C3_F3_T3.width
        sheet.cell(row=4,column=count).value = Burton_C3_F3_T3.length
        for j in range(8, sheet.max_row+1):
                           
            if Burton_C3_F3_T3.width <= sheet2.cell(row=j,column=2).value and Burton_C3_F3_T3.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Burton_C3_F3_T3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Burton C, F, & T w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Burton C, F, & T w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Burton_C3_F3_T3.width
        sheet.cell(row=4,column=count).value = Burton_C3_F3_T3.cp
        for j in range(8, sheet.max_row+1):
                           
            if Burton_C3_F3_T3.width <= sheet2.cell(row=j,column=2).value and Burton_C3_F3_T3.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Burton_C3_F3_T3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Burton C, F, & T w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Centennial A, B, & C w/ Patio"):
        sheet.cell(row=3,column=count).value = Centennial_A_B_C.width
        sheet.cell(row=4,column=count).value = Centennial_A_B_C.length
        for j in range(8, sheet.max_row+1):
                           
            if Centennial_A_B_C.width <= sheet2.cell(row=j,column=2).value and Centennial_A_B_C.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Centennial_A_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Centennial A, B, & C w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")

    elif (sheet.cell(row=7,column=count).value == "Centennial A, B, & C w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Centennial_A_B_C.width
        sheet.cell(row=4,column=count).value = Centennial_A_B_C.cp
        for j in range(8, sheet.max_row+1):
                           
            if Centennial_A_B_C.width <= sheet2.cell(row=j,column=2).value and Centennial_A_B_C.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Centennial_A_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Centennial A, B, & C w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Crestwood A, B, C, D, E, & F w/ Patio"):
        sheet.cell(row=3,column=count).value = Crestwood_A_B_C_D_E_F.width
        sheet.cell(row=4,column=count).value = Crestwood_A_B_C_D_E_F.length
        for j in range(8, sheet.max_row+1):
                           
            if Crestwood_A_B_C_D_E_F.width <= sheet2.cell(row=j,column=2).value and Crestwood_A_B_C_D_E_F.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Crestwood_A_B_C_D_E_F.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Crestwood A, B, C, D, E, & F w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
    
    elif (sheet.cell(row=7,column=count).value == "Cumberland A, B, & C w/ Patio") or (sheet.cell(row=7,column=count).value == "Cumberland A, B, & C w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Cumberland_A_B_C.width
        sheet.cell(row=4,column=count).value = Cumberland_A_B_C.length
        for j in range(8, sheet.max_row+1):
                           
            if Cumberland_A_B_C.width <= sheet2.cell(row=j,column=2).value and Cumberland_A_B_C.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Cumberland_A_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Cumberland A, B, & C w/ Patio and Cumberland A, B, & C w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Cumberland A, B, & C w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Cumberland A, B, & C w/ Ext. Covered Patio"):
        sheet.cell(row=3,column=count).value = Cumberland_A_B_C.width
        sheet.cell(row=4,column=count).value = Cumberland_A_B_C.cp
        for j in range(8, sheet.max_row+1):
                           
            if Cumberland_A_B_C.width <= sheet2.cell(row=j,column=2).value and Cumberland_A_B_C.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Cumberland_A_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Cumberland A, B, & C w/ Covered Patio and Cumberland A, B, & C w/ Ext. Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Dakota A, B, & E w/ Patio") or (sheet.cell(row=7,column=count).value == "Dakota A, B, & E w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Dakota_A_B_E.width
        sheet.cell(row=4,column=count).value = Dakota_A_B_E.length
        for j in range(8, sheet.max_row+1):
                           
            if Dakota_A_B_E.width <= sheet2.cell(row=j,column=2).value and Dakota_A_B_E.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Dakota_A_B_E.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Dakota A, B, & E w/ Patio and Dakota A, B, & E w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Dakota A, B, & E w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Dakota_A_B_E.width
        sheet.cell(row=4,column=count).value = Dakota_A_B_E.cp
        for j in range(8, sheet.max_row+1):
                           
            if Dakota_A_B_E.width <= sheet2.cell(row=j,column=2).value and Dakota_A_B_E.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Dakota_A_B_E.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Dakota A, B, & E w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Dakota C & D w/ Patio") or (sheet.cell(row=7,column=count).value == "Dakota C & D w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Dakota_C_D.width
        sheet.cell(row=4,column=count).value = Dakota_C_D.length
        for j in range(8, sheet.max_row+1):
                           
            if Dakota_C_D.width <= sheet2.cell(row=j,column=2).value and Dakota_C_D.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Dakota_C_D.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Dakota C & D w/ Patio and Dakota C & D w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Dakota C & D w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Dakota_C_D.width
        sheet.cell(row=4,column=count).value = Dakota_C_D.cp
        for j in range(8, sheet.max_row+1):
                           
            if Dakota_C_D.width <= sheet2.cell(row=j,column=2).value and Dakota_C_D.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Dakota_C_D.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Dakota C & D w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Dakota A, B, & E Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Dakota A, B, & E Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Dakota_A_B_E.width
        sheet.cell(row=4,column=count).value = Dakota_A_B_E.length
        for j in range(8, sheet.max_row+1):
                           
            if Dakota_A_B_E.width <= sheet2.cell(row=j,column=2).value and Dakota_A_B_E.length <= sheet2.cell(row=j,column=3).value and Dakota_A_B_E.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Dakota_A_B_E.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Dakota_A_B_E.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Dakota A, B, & E Side Entry w/ Patio and Dakota A, B, & E Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Dakota A, B, & E Side Entry w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Dakota_A_B_E.width
        sheet.cell(row=4,column=count).value = Dakota_A_B_E.cp
        for j in range(8, sheet.max_row+1):
                           
            if Dakota_A_B_E.width <= sheet2.cell(row=j,column=2).value and Dakota_A_B_E.cp <= sheet2.cell(row=j,column=3).value and Dakota_A_B_E.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Dakota_A_B_E.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Dakota_A_B_E.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Dakota A, B, & E Side Entry w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Dakota C & D Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Dakota C & D Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Dakota_C_D.width
        sheet.cell(row=4,column=count).value = Dakota_C_D.length
        for j in range(8, sheet.max_row+1):
                           
            if Dakota_C_D.width <= sheet2.cell(row=j,column=2).value and Dakota_C_D.length <= sheet2.cell(row=j,column=3).value and Dakota_C_D.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Dakota_C_D.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Dakota_C_D.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Dakota C & D Side Entry w/ Patio and Dakota C & D Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Dakota C & D Side Entry w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Dakota_C_D.width
        sheet.cell(row=4,column=count).value = Dakota_C_D.cp
        for j in range(8, sheet.max_row+1):
                           
            if Dakota_C_D.width <= sheet2.cell(row=j,column=2).value and Dakota_C_D.cp <= sheet2.cell(row=j,column=3).value and Dakota_C_D.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Dakota_C_D.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Dakota_C_D.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Dakota C & D Side Entry w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Firethorn A, B, C, M, N, & O w/ Patio") or (sheet.cell(row=7,column=count).value == "Firethorn A, B, C, M, N, & O w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Firethorn_A_B_C_M_N_O.width
        sheet.cell(row=4,column=count).value = Firethorn_A_B_C_M_N_O.length
        for j in range(8, sheet.max_row+1):
                           
            if Firethorn_A_B_C_M_N_O.width <= sheet2.cell(row=j,column=2).value and Firethorn_A_B_C_M_N_O.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Firethorn_A_B_C_M_N_O.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Firethorn A, B, C, M, N, & O w/ Patio and Firethorn A, B, C, M, N, & O w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Firethorn A, B, C, M, N, & O Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Firethorn A, B, C, M, N, & O Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Firethorn_A_B_C_M_N_O.width
        sheet.cell(row=4,column=count).value = Firethorn_A_B_C_M_N_O.length
        for j in range(8, sheet.max_row+1):
                           
            if Firethorn_A_B_C_M_N_O.width <= sheet2.cell(row=j,column=2).value and Firethorn_A_B_C_M_N_O.length <= sheet2.cell(row=j,column=3).value and Firethorn_A_B_C_M_N_O.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Firethorn_A_B_C_M_N_O.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Firethorn_A_B_C_M_N_O.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Firethorn A, B, C, M, N, & O Side Entry w/ Patio and Firethorn A, B, C, M, N, & O Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Fontana C, F, & T w/ Patio") or (sheet.cell(row=7,column=count).value == "Fontana C, F, & T w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Fontana_C3_F3_T3.width
        sheet.cell(row=4,column=count).value = Fontana_C3_F3_T3.length
        for j in range(8, sheet.max_row+1):
                           
            if Fontana_C3_F3_T3.width <= sheet2.cell(row=j,column=2).value and Fontana_C3_F3_T3.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Fontana_C3_F3_T3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Fontana C, F, & T w/ Patio and Fontana C, F, & T w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Fontana C, F, & T w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Fontana_C3_F3_T3.width
        sheet.cell(row=4,column=count).value = Fontana_C3_F3_T3.cp
        for j in range(8, sheet.max_row+1):
                           
            if Fontana_C3_F3_T3.width <= sheet2.cell(row=j,column=2).value and Fontana_C3_F3_T3.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Fontana_C3_F3_T3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Fontana C, F, & T w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Fontana C, F, & T w/ Ext. Covered Patio"):
        sheet.cell(row=3,column=count).value = Fontana_C3_F3_T3_ext.width
        sheet.cell(row=4,column=count).value = Fontana_C3_F3_T3_ext.cp
        for j in range(8, sheet.max_row+1):
                           
            if Fontana_C3_F3_T3_ext.width <= sheet2.cell(row=j,column=2).value and Fontana_C3_F3_T3_ext.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Fontana_C3_F3_T3_ext.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Fontana C, F, & T w/ Ext. Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Greenfield C, F, & T w/ Patio") or (sheet.cell(row=7,column=count).value == "Greenfield C, F, & T w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Greenfield_C3_F3_T3.width
        sheet.cell(row=4,column=count).value = Greenfield_C3_F3_T3.length
        for j in range(8, sheet.max_row+1):
                           
            if Greenfield_C3_F3_T3.width <= sheet2.cell(row=j,column=2).value and Greenfield_C3_F3_T3.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Greenfield_C3_F3_T3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Greenfield C, F, & T w/ Patio and Greenfield C, F, & T w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Greenfield C, F, & T w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Greenfield_C3_F3_T3.width
        sheet.cell(row=4,column=count).value = Greenfield_C3_F3_T3.cp
        for j in range(8, sheet.max_row+1):
                           
            if Greenfield_C3_F3_T3.width <= sheet2.cell(row=j,column=2).value and Greenfield_C3_F3_T3.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Greenfield_C3_F3_T3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Greenfield C, F, & T w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Greenfield C, F, & T w/ Ext. Covered Patio"):
        sheet.cell(row=3,column=count).value = Greenfield_C3_F3_T3_ext.width
        sheet.cell(row=4,column=count).value = Greenfield_C3_F3_T3_ext.cp
        for j in range(8, sheet.max_row+1):
                           
            if Greenfield_C3_F3_T3_ext.width <= sheet2.cell(row=j,column=2).value and Greenfield_C3_F3_T3_ext.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Greenfield_C3_F3_T3_ext.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Greenfield C, F, & T w/ Ext. Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Harding C, F, & T w/ Patio") or (sheet.cell(row=7,column=count).value == "Harding C, F, & T w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Harding_C3_F3_T3.width
        sheet.cell(row=4,column=count).value = Harding_C3_F3_T3.length
        for j in range(8, sheet.max_row+1):
                           
            if Harding_C3_F3_T3.width <= sheet2.cell(row=j,column=2).value and Harding_C3_F3_T3.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Harding_C3_F3_T3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Harding C, F, & T w/ Patio and Harding C, F, & T w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Harding C, F, & T w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Harding_C3_F3_T3.width
        sheet.cell(row=4,column=count).value = Harding_C3_F3_T3.cp
        for j in range(8, sheet.max_row+1):
                           
            if Harding_C3_F3_T3.width <= sheet2.cell(row=j,column=2).value and Harding_C3_F3_T3.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Harding_C3_F3_T3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Harding C, F, & T w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Harding C, F, & T w/ Ext. Covered Patio"):
        sheet.cell(row=3,column=count).value = Harding_C3_F3_T3_ext.width
        sheet.cell(row=4,column=count).value = Harding_C3_F3_T3_ext.cp
        for j in range(8, sheet.max_row+1):
                           
            if Harding_C3_F3_T3_ext.width <= sheet2.cell(row=j,column=2).value and Harding_C3_F3_T3_ext.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Harding_C3_F3_T3_ext.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Harding C, F, & T w/ Ext. Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Jordan A, B, & C w/ Patio"):
        sheet.cell(row=3,column=count).value = Jordan_A1_B2_C3.width
        sheet.cell(row=4,column=count).value = Jordan_A1_B2_C3.length
        for j in range(8, sheet.max_row+1):
                           
            if Jordan_A1_B2_C3.width <= sheet2.cell(row=j,column=2).value and Jordan_A1_B2_C3.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Jordan_A1_B2_C3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Jordan A, B, & C w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Jordan A, B, & C w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Jordan_A1_B2_C3.width
        sheet.cell(row=4,column=count).value = Jordan_A1_B2_C3.cp
        for j in range(8, sheet.max_row+1):
                           
            if Jordan_A1_B2_C3.width <= sheet2.cell(row=j,column=2).value and Jordan_A1_B2_C3.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Jordan_A1_B2_C3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Jordan A, B, & C w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E w/ Patio") or (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Juniper_A_B_D_E.width
        sheet.cell(row=4,column=count).value = Juniper_A_B_D_E.length
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_A_B_D_E.width <= sheet2.cell(row=j,column=2).value and Juniper_A_B_D_E.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_A_B_D_E.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper A, B, D, & E w/ Patio and Juniper A, B, D, & E w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Juniper_A_B_D_E.width
        sheet.cell(row=4,column=count).value = Juniper_A_B_D_E.cp
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_A_B_D_E.width <= sheet2.cell(row=j,column=2).value and Juniper_A_B_D_E.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_A_B_D_E.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper A, B, D, & E w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper C w/ Patio") or (sheet.cell(row=7,column=count).value == "Juniper C w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Juniper_C.width
        sheet.cell(row=4,column=count).value = Juniper_C.length
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_C.width <= sheet2.cell(row=j,column=2).value and Juniper_C.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper C w/ Patio and Juniper C w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper C w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Juniper_C.width
        sheet.cell(row=4,column=count).value = Juniper_C.cp
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_C.width <= sheet2.cell(row=j,column=2).value and Juniper_C.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper C w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E 3 Front w/ Patio") or (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E 3 Front w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Juniper_A_B_D_E_3_Front.width
        sheet.cell(row=4,column=count).value = Juniper_A_B_D_E_3_Front.length
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_A_B_D_E_3_Front.width <= sheet2.cell(row=j,column=2).value and Juniper_A_B_D_E_3_Front.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_A_B_D_E_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper A, B, D, & E 3 Front w/ Patio and Juniper A, B, D, & E 3 Front w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E 3 Front w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Juniper_A_B_D_E_3_Front.width
        sheet.cell(row=4,column=count).value = Juniper_A_B_D_E_3_Front.cp
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_A_B_D_E_3_Front.width <= sheet2.cell(row=j,column=2).value and Juniper_A_B_D_E_3_Front.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_A_B_D_E_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper A, B, D, & E 3 Front w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper C 3 Front w/ Patio") or (sheet.cell(row=7,column=count).value == "Juniper C 3 Front w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Juniper_C_3_Front.width
        sheet.cell(row=4,column=count).value = Juniper_C_3_Front.length
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_C_3_Front.width <= sheet2.cell(row=j,column=2).value and Juniper_C_3_Front.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_C_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper C 3 Front w/ Patio and Juniper C 3 Front w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper C 3 Front w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Juniper_C_3_Front.width
        sheet.cell(row=4,column=count).value = Juniper_C_3_Front.cp
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_C_3_Front.width <= sheet2.cell(row=j,column=2).value and Juniper_C_3_Front.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_C_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper C 3 Front w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Juniper_A_B_D_E.width
        sheet.cell(row=4,column=count).value = Juniper_A_B_D_E.length
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_A_B_D_E.width <= sheet2.cell(row=j,column=2).value and Juniper_A_B_D_E.length <= sheet2.cell(row=j,column=3).value and Juniper_A_B_D_E.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Juniper_A_B_D_E.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_A_B_D_E.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper A, B, D, & E Side Entry w/ Patio and Juniper A, B, D, & E Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper A, B, D, & E Side Entry w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Juniper_A_B_D_E.width
        sheet.cell(row=4,column=count).value = Juniper_A_B_D_E.cp
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_A_B_D_E.width <= sheet2.cell(row=j,column=2).value and Juniper_A_B_D_E.cp <= sheet2.cell(row=j,column=3).value and Juniper_A_B_D_E.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Juniper_A_B_D_E.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_A_B_D_E.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper A, B, D, & E Side Entry w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                            
    elif (sheet.cell(row=7,column=count).value == "Juniper C Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Juniper C Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Juniper_C.width
        sheet.cell(row=4,column=count).value = Juniper_C.length
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_C.width <= sheet2.cell(row=j,column=2).value and Juniper_C.length <= sheet2.cell(row=j,column=3).value and Juniper_C.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Juniper_C.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper C Side Entry w/ Patio and Juniper C Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Juniper C Side Entry w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Juniper_C.width
        sheet.cell(row=4,column=count).value = Juniper_C.cp
        for j in range(8, sheet.max_row+1):
                           
            if Juniper_C.width <= sheet2.cell(row=j,column=2).value and Juniper_C.cp <= sheet2.cell(row=j,column=3).value and Juniper_C.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Juniper_C.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Juniper_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Juniper C Side Entry w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Mcdowell A w/ Patio"):
        sheet.cell(row=3,column=count).value = Mcdowell_A2.width
        sheet.cell(row=4,column=count).value = Mcdowell_A2.length
        for j in range(8, sheet.max_row+1):
                           
            if Mcdowell_A2.width <= sheet2.cell(row=j,column=2).value and Mcdowell_A2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Mcdowell_A2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Mcdowell A w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Mcdowell A w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Mcdowell_A2.width
        sheet.cell(row=4,column=count).value = Mcdowell_A2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Mcdowell_A2.width <= sheet2.cell(row=j,column=2).value and Mcdowell_A2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Mcdowell_A2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Mcdowell A w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Mcdowell B & C w/ Patio"):
        sheet.cell(row=3,column=count).value = Mcdowell_B2_C2.width
        sheet.cell(row=4,column=count).value = Mcdowell_B2_C2.length
        for j in range(8, sheet.max_row+1):
                           
            if Mcdowell_B2_C2.width <= sheet2.cell(row=j,column=2).value and Mcdowell_B2_C2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Mcdowell_B2_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Mcdowell B & C w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Mcdowell B & C w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Mcdowell_B2_C2.width
        sheet.cell(row=4,column=count).value = Mcdowell_B2_C2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Mcdowell_B2_C2.width <= sheet2.cell(row=j,column=2).value and Mcdowell_B2_C2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Mcdowell_B2_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Mcdowell B & C w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Middleton A & D w/ Patio") or (sheet.cell(row=7,column=count).value == "Middleton A & D w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Middleton_A2_D2.width
        sheet.cell(row=4,column=count).value = Middleton_A2_D2.length
        for j in range(8, sheet.max_row+1):
                           
            if Middleton_A2_D2.width <= sheet2.cell(row=j,column=2).value and Middleton_A2_D2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Middleton_A2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Middleton A & D w/ Patio and Middleton A & D w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Middleton A & D w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Middleton A & D w/ Ext. Covered Patio"):
        sheet.cell(row=3,column=count).value = Middleton_A2_D2.width
        sheet.cell(row=4,column=count).value = Middleton_A2_D2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Middleton_A2_D2.width <= sheet2.cell(row=j,column=2).value and Middleton_A2_D2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Middleton_A2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Middleton A & D w/ Covered Patio and Middleton A & D w/ Ext. Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")

                
    elif (sheet.cell(row=7,column=count).value == "Middleton A & D Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Middleton A & D Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Middleton_A2_D2.width
        sheet.cell(row=4,column=count).value = Middleton_A2_D2.length
        for j in range(8, sheet.max_row+1):
                           
            if Middleton_A2_D2.width <= sheet2.cell(row=j,column=2).value and Middleton_A2_D2.length <= sheet2.cell(row=j,column=3).value and Middleton_A2_D2.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Middleton_A2_D2.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Middleton_A2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Middleton A & D Side Entry w/ Patio and Middleton A & D Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Middleton A & D Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Middleton A & D Side Entry w/ Covered Ext. Patio"):
        sheet.cell(row=3,column=count).value = Middleton_A2_D2.width
        sheet.cell(row=4,column=count).value = Middleton_A2_D2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Middleton_A2_D2.width <= sheet2.cell(row=j,column=2).value and Middleton_A2_D2.cp <= sheet2.cell(row=j,column=3).value and Middleton_A2_D2.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Middleton_A2_D2.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Middleton_A2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Middleton A & D Side Entry w/ Covered Patio and Middleton A & D Side Entry w/ Covered Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Middleton B & C w/ Patio") or (sheet.cell(row=7,column=count).value == "Middleton B & C w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Middleton_B2_C2.width
        sheet.cell(row=4,column=count).value = Middleton_B2_C2.length
        for j in range(8, sheet.max_row+1):
                           
            if Middleton_B2_C2.width <= sheet2.cell(row=j,column=2).value and Middleton_B2_C2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Middleton_B2_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Middleton B & C w/ Patio and Middleton B & C w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Middleton B & C w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Middleton B & C w/ Ext. Covered Patio"):
        sheet.cell(row=3,column=count).value = Middleton_B2_C2.width
        sheet.cell(row=4,column=count).value = Middleton_B2_C2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Middleton_B2_C2.width <= sheet2.cell(row=j,column=2).value and Middleton_B2_C2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Middleton_B2_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Middleton B & C w/ Covered Patio and Middleton B & C w/ Ext. Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Middleton B & C Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Middleton B & C Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Middleton_B2_C2.width
        sheet.cell(row=4,column=count).value = Middleton_B2_C2.length
        for j in range(8, sheet.max_row+1):
                           
            if Middleton_B2_C2.width <= sheet2.cell(row=j,column=2).value and Middleton_B2_C2.length <= sheet2.cell(row=j,column=3).value and Middleton_B2_C2.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Middleton_B2_C2.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Middleton_B2_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Middleton B & C Side Entry w/ Patio and Middleton B & C Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Middleton B & C Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Middleton B & C Side Entry w/ Covered Ext. Patio"):
        sheet.cell(row=3,column=count).value = Middleton_B2_C2.width
        sheet.cell(row=4,column=count).value = Middleton_B2_C2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Middleton_B2_C2.width <= sheet2.cell(row=j,column=2).value and Middleton_B2_C2.cp <= sheet2.cell(row=j,column=3).value and Middleton_B2_C2.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Middleton_B2_C2.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Middleton_B2_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Middleton B & C Side Entry w/ Covered Patio and Middleton B & C Side Entry w/ Covered Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina A & M w/ Patio"):
        sheet.cell(row=3,column=count).value = Nandina_A_M.width
        sheet.cell(row=4,column=count).value = Nandina_A_M.length
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_A_M.width <= sheet2.cell(row=j,column=2).value and Nandina_A_M.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_A_M.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina A & M w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina A & M w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Nandina_A_M.width
        sheet.cell(row=4,column=count).value = Nandina_A_M.cp
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_A_M.width <= sheet2.cell(row=j,column=2).value and Nandina_A_M.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_A_M.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina A & M w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina B, C, N, & O w/ Patio"):
        sheet.cell(row=3,column=count).value = Nandina_B_C_N_O.width
        sheet.cell(row=4,column=count).value = Nandina_B_C_N_O.length
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_B_C_N_O.width <= sheet2.cell(row=j,column=2).value and Nandina_B_C_N_O.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_B_C_N_O.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina B, C, N, & O w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina B, C, N, & O w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Nandina_B_C_N_O.width
        sheet.cell(row=4,column=count).value = Nandina_B_C_N_O.cp
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_B_C_N_O.width <= sheet2.cell(row=j,column=2).value and Nandina_B_C_N_O.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_B_C_N_O.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina B, C, N, & O w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina A & M 3 Front w/ Patio"):
        sheet.cell(row=3,column=count).value = Nandina_A_M_3_Front.width
        sheet.cell(row=4,column=count).value = Nandina_A_M_3_Front.length
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_A_M_3_Front.width <= sheet2.cell(row=j,column=2).value and Nandina_A_M_3_Front.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_A_M_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina A & M 3 Front w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina A & M 3 Front w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Nandina_A_M_3_Front.width
        sheet.cell(row=4,column=count).value = Nandina_A_M_3_Front.cp
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_A_M_3_Front.width <= sheet2.cell(row=j,column=2).value and Nandina_A_M_3_Front.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_A_M_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina A & M 3 Front w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina B, C, N, & O 3 Front w/ Patio"):
        sheet.cell(row=3,column=count).value = Nandina_B_C_N_O_3_Front.width
        sheet.cell(row=4,column=count).value = Nandina_B_C_N_O_3_Front.length
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_B_C_N_O_3_Front.width <= sheet2.cell(row=j,column=2).value and Nandina_B_C_N_O_3_Front.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_B_C_N_O_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina B, C, N, & O 3 Front w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina B, C, N, & O 3 Front w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Nandina_B_C_N_O_3_Front.width
        sheet.cell(row=4,column=count).value = Nandina_B_C_N_O_3_Front.cp
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_B_C_N_O_3_Front.width <= sheet2.cell(row=j,column=2).value and Nandina_B_C_N_O_3_Front.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_B_C_N_O_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina B, C, N, & O 3 Front w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina A & M Side Entry w/ Patio"):
        sheet.cell(row=3,column=count).value = Nandina_A_M_Side.width
        sheet.cell(row=4,column=count).value = Nandina_A_M_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_A_M_Side.width <= sheet2.cell(row=j,column=2).value and Nandina_A_M_Side.length <= sheet2.cell(row=j,column=3).value and Nandina_A_M_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Nandina_A_M_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_A_M_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina A & M Side Entry w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina A & M Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Nandina_A_M_Side.width
        sheet.cell(row=4,column=count).value = Nandina_A_M_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_A_M_Side.width <= sheet2.cell(row=j,column=2).value and Nandina_A_M_Side.cp <= sheet2.cell(row=j,column=3).value and Nandina_A_M_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Nandina_A_M_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_A_M_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina A & M Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina B, C, N, & O Side Entry w/ Patio"):
        sheet.cell(row=3,column=count).value = Nandina_B_C_N_O_Side.width
        sheet.cell(row=4,column=count).value = Nandina_B_C_N_O_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_B_C_N_O_Side.width <= sheet2.cell(row=j,column=2).value and Nandina_B_C_N_O_Side.length <= sheet2.cell(row=j,column=3).value and Nandina_B_C_N_O_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Nandina_B_C_N_O_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_B_C_N_O_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina B, C, N, & O Side Entry w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina B, C, N, & O Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Nandina_B_C_N_O_Side.width
        sheet.cell(row=4,column=count).value = Nandina_B_C_N_O_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_B_C_N_O_Side.width <= sheet2.cell(row=j,column=2).value and Nandina_B_C_N_O_Side.cp <= sheet2.cell(row=j,column=3).value and Nandina_B_C_N_O_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Nandina_B_C_N_O_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_B_C_N_O_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina B, C, N, & O Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina A & M 3 Bolt On w/ Patio"):
        sheet.cell(row=3,column=count).value = Nandina_A_M_3_BoltOn.width
        sheet.cell(row=4,column=count).value = Nandina_A_M_3_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_A_M_3_BoltOn.width <= sheet2.cell(row=j,column=2).value and Nandina_A_M_3_BoltOn.length <= sheet2.cell(row=j,column=3).value and Nandina_A_M_3_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Nandina_A_M_3_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_A_M_3_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina A & M 3 Bolt On w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina A & M 3 Bolt On w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Nandina_A_M_3_BoltOn.width
        sheet.cell(row=4,column=count).value = Nandina_A_M_3_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_A_M_3_BoltOn.width <= sheet2.cell(row=j,column=2).value and Nandina_A_M_3_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Nandina_A_M_3_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Nandina_A_M_3_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_A_M_3_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina A & M 3 Bolt On w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina B, C, N, & O 3 Bolt On w/ Patio"):
        sheet.cell(row=3,column=count).value = Nandina_B_C_N_O_3_BoltOn.width
        sheet.cell(row=4,column=count).value = Nandina_B_C_N_O_3_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_B_C_N_O_3_BoltOn.width <= sheet2.cell(row=j,column=2).value and Nandina_B_C_N_O_3_BoltOn.length <= sheet2.cell(row=j,column=3).value and Nandina_B_C_N_O_3_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Nandina_B_C_N_O_3_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_B_C_N_O_3_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina B, C, N, & O 3 Bolt On w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Nandina B, C, N, & O 3 Bolt On w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Nandina_B_C_N_O_3_BoltOn.width
        sheet.cell(row=4,column=count).value = Nandina_B_C_N_O_3_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Nandina_B_C_N_O_3_BoltOn.width <= sheet2.cell(row=j,column=2).value and Nandina_B_C_N_O_3_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Nandina_B_C_N_O_3_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Nandina_B_C_N_O_3_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Nandina_B_C_N_O_3_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Nandina B, C, N, & O 3 Bolt On w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte A w/ Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_A.width
        sheet.cell(row=4,column=count).value = Oakmonte_A.length
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_A.width <= sheet2.cell(row=j,column=2).value and Oakmonte_A.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_A.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte A w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte A w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_A.width
        sheet.cell(row=4,column=count).value = Oakmonte_A.cp
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_A.width <= sheet2.cell(row=j,column=2).value and Oakmonte_A.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_A.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte A w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte B, C, D, E, F, G, & H w/ Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_B_C_D_E_F_G_H.width
        sheet.cell(row=4,column=count).value = Oakmonte_B_C_D_E_F_G_H.length
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_B_C_D_E_F_G_H.width <= sheet2.cell(row=j,column=2).value and Oakmonte_B_C_D_E_F_G_H.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_B_C_D_E_F_G_H.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte B, C, D, E, F, G, & H w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte B, C, D, E, F, G, & H w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_B_C_D_E_F_G_H.width
        sheet.cell(row=4,column=count).value = Oakmonte_B_C_D_E_F_G_H.cp
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_B_C_D_E_F_G_H.width <= sheet2.cell(row=j,column=2).value and Oakmonte_B_C_D_E_F_G_H.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_B_C_D_E_F_G_H.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte B, C, D, E, F, G, & H w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte A Side Entry w/ Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_A_Side.width
        sheet.cell(row=4,column=count).value = Oakmonte_A_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_A_Side.width <= sheet2.cell(row=j,column=2).value and Oakmonte_A_Side.length <= sheet2.cell(row=j,column=3).value and Oakmonte_A_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Oakmonte_A_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_A_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte A Side Entry w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte A Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_A_Side.width
        sheet.cell(row=4,column=count).value = Oakmonte_A_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_A_Side.width <= sheet2.cell(row=j,column=2).value and Oakmonte_A_Side.cp <= sheet2.cell(row=j,column=3).value and Oakmonte_A_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Oakmonte_A_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_A_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte A Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte B, C, D, E, F, G, & H Side Entry w/ Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_B_C_D_E_F_G_H_Side.width
        sheet.cell(row=4,column=count).value = Oakmonte_B_C_D_E_F_G_H_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_B_C_D_E_F_G_H_Side.width <= sheet2.cell(row=j,column=2).value and Oakmonte_B_C_D_E_F_G_H_Side.length <= sheet2.cell(row=j,column=3).value and Oakmonte_B_C_D_E_F_G_H_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Oakmonte_B_C_D_E_F_G_H_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_B_C_D_E_F_G_H_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte B, C, D, E, F, G, & H Side Entry w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte B, C, D, E, F, G, & H Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_B_C_D_E_F_G_H_Side.width
        sheet.cell(row=4,column=count).value = Oakmonte_B_C_D_E_F_G_H_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_B_C_D_E_F_G_H_Side.width <= sheet2.cell(row=j,column=2).value and Oakmonte_B_C_D_E_F_G_H_Side.cp <= sheet2.cell(row=j,column=3).value and Oakmonte_B_C_D_E_F_G_H_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Oakmonte_B_C_D_E_F_G_H_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_B_C_D_E_F_G_H_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte B, C, D, E, F, G, & H Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte A Bolt On w/ Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_A_BoltOn.width
        sheet.cell(row=4,column=count).value = Oakmonte_A_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_A_BoltOn.width <= sheet2.cell(row=j,column=2).value and Oakmonte_A_BoltOn.length <= sheet2.cell(row=j,column=3).value and Oakmonte_A_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Oakmonte_A_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_A_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte A Bolt On w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte A Bolt On w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_A_BoltOn.width
        sheet.cell(row=4,column=count).value = Oakmonte_A_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_A_BoltOn.width <= sheet2.cell(row=j,column=2).value and Oakmonte_A_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Oakmonte_A_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Oakmonte_A_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_A_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte A Bolt On w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte B, D, E, F, G, & H Bolt On w/ Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_B_C_D_E_F_G_H_BoltOn.width
        sheet.cell(row=4,column=count).value = Oakmonte_B_C_D_E_F_G_H_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_B_C_D_E_F_G_H_BoltOn.width <= sheet2.cell(row=j,column=2).value and Oakmonte_B_C_D_E_F_G_H_BoltOn.length <= sheet2.cell(row=j,column=3).value and Oakmonte_B_C_D_E_F_G_H_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Oakmonte_B_C_D_E_F_G_H_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_B_C_D_E_F_G_H_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte B, D, E, F, G, & H Bolt On w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Oakmonte B, D, E, F, G, & H Bolt On w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Oakmonte_B_C_D_E_F_G_H_BoltOn.width
        sheet.cell(row=4,column=count).value = Oakmonte_B_C_D_E_F_G_H_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Oakmonte_B_C_D_E_F_G_H_BoltOn.width <= sheet2.cell(row=j,column=2).value and Oakmonte_B_C_D_E_F_G_H_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Oakmonte_B_C_D_E_F_G_H_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Oakmonte_B_C_D_E_F_G_H_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Oakmonte_B_C_D_E_F_G_H_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Oakmonte B, D, E, F, G, & H Bolt On w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds A w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_A2.width
        sheet.cell(row=4,column=count).value = Reynolds_A2.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_A2.width <= sheet2.cell(row=j,column=2).value and Reynolds_A2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_A2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds A w/ Patio and Reynolds A w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds A w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds A w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_A2.width
        sheet.cell(row=4,column=count).value = Reynolds_A2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_A2.width <= sheet2.cell(row=j,column=2).value and Reynolds_A2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_A2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds A w/ Covered Patio, Reynolds A w/ Ext. Covered Patio, Reynolds A w/ Morning Room, and Reynolds A w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds B w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_B2.width
        sheet.cell(row=4,column=count).value = Reynolds_B2.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_B2.width <= sheet2.cell(row=j,column=2).value and Reynolds_B2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_B2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds B w/ Patio and Reynolds B w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds B w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds B w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_B2.width
        sheet.cell(row=4,column=count).value = Reynolds_B2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_B2.width <= sheet2.cell(row=j,column=2).value and Reynolds_B2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_B2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds B w/ Covered Patio, Reynolds B w/ Ext. Covered Patio, Reynolds B w/ Morning Room, and Reynolds B w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds C w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_C2.width
        sheet.cell(row=4,column=count).value = Reynolds_C2.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_C2.width <= sheet2.cell(row=j,column=2).value and Reynolds_C2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds C w/ Patio and Reynolds C w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds C w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds C w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_C2.width
        sheet.cell(row=4,column=count).value = Reynolds_C2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_C2.width <= sheet2.cell(row=j,column=2).value and Reynolds_C2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_C2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds C w/ Covered Patio, Reynolds C w/ Ext. Covered Patio, Reynolds C w/ Morning Room, and Reynolds C w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
           
    elif (sheet.cell(row=7,column=count).value == "Reynolds D w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_D2.width
        sheet.cell(row=4,column=count).value = Reynolds_D2.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_D2.width <= sheet2.cell(row=j,column=2).value and Reynolds_D2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds D w/ Patio and Reynolds D w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds D w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds D w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_D2.width
        sheet.cell(row=4,column=count).value = Reynolds_D2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_D2.width <= sheet2.cell(row=j,column=2).value and Reynolds_D2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds D w/ Covered Patio, Reynolds D w/ Ext. Covered Patio, Reynolds D w/ Morning Room, and Reynolds D w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds A Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_A2_Side.width
        sheet.cell(row=4,column=count).value = Reynolds_A2_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_A2_Side.width <= sheet2.cell(row=j,column=2).value and Reynolds_A2_Side.length <= sheet2.cell(row=j,column=3).value and Reynolds_A2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_A2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_A2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds A Side Entry w/ Patio and Reynolds A Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds A Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds A Side Entry w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_A2_Side.width
        sheet.cell(row=4,column=count).value = Reynolds_A2_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_A2_Side.width <= sheet2.cell(row=j,column=2).value and Reynolds_A2_Side.cp <= sheet2.cell(row=j,column=3).value and Reynolds_A2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_A2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_A2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds A Side Entry w/ Covered Patio, Reynolds A Side Entry w/ Ext. Covered Patio, Reynolds A Side Entry w/ Morning Room, and Reynolds A Side Entry w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds B Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_B2_Side.width
        sheet.cell(row=4,column=count).value = Reynolds_B2_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_B2_Side.width <= sheet2.cell(row=j,column=2).value and Reynolds_B2_Side.length <= sheet2.cell(row=j,column=3).value and Reynolds_B2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_B2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_B2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds B Side Entry w/ Patio and Reynolds B Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds B Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds B Side Entry w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_B2_Side.width
        sheet.cell(row=4,column=count).value = Reynolds_B2_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_B2_Side.width <= sheet2.cell(row=j,column=2).value and Reynolds_B2_Side.cp <= sheet2.cell(row=j,column=3).value and Reynolds_B2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_B2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_B2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds B Side Entry w/ Covered Patio, Reynolds B Side Entry w/ Ext. Covered Patio, Reynolds B Side Entry w/ Morning Room, and Reynolds B Side Entry w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds C Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_C2_Side.width
        sheet.cell(row=4,column=count).value = Reynolds_C2_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_C2_Side.width <= sheet2.cell(row=j,column=2).value and Reynolds_C2_Side.length <= sheet2.cell(row=j,column=3).value and Reynolds_C2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_C2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_C2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds C Side Entry w/ Patio and Reynolds C Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds C Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds C Side Entry w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_C2_Side.width
        sheet.cell(row=4,column=count).value = Reynolds_C2_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_C2_Side.width <= sheet2.cell(row=j,column=2).value and Reynolds_C2_Side.cp <= sheet2.cell(row=j,column=3).value and Reynolds_C2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_C2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_C2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds C Side Entry w/ Covered Patio, Reynolds C Side Entry w/ Ext. Covered Patio, Reynolds C Side Entry w/ Morning Room, and Reynolds C Side Entry w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds D Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_D2_Side.width
        sheet.cell(row=4,column=count).value = Reynolds_D2_Side.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_D2_Side.width <= sheet2.cell(row=j,column=2).value and Reynolds_D2_Side.length <= sheet2.cell(row=j,column=3).value and Reynolds_D2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_D2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_D2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds D Side Entry w/ Patio and Reynolds D Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds D Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds D Side Entry w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_D2_Side.width
        sheet.cell(row=4,column=count).value = Reynolds_D2_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_D2_Side.width <= sheet2.cell(row=j,column=2).value and Reynolds_D2_Side.cp <= sheet2.cell(row=j,column=3).value and Reynolds_D2_Side.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_D2_Side.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_D2_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds D Side Entry w/ Patio and Reynolds D Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds A Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_A2_BoltOn.width
        sheet.cell(row=4,column=count).value = Reynolds_A2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_A2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Reynolds_A2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Reynolds_A2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_A2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_A2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds A Bolt On w/ Patio and Reynolds A Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds A Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds A Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds A Bolt On w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_A2_BoltOn.width
        sheet.cell(row=4,column=count).value = Reynolds_A2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_A2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Reynolds_A2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Reynolds_A2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_A2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_A2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds A Bolt On w/ Covered Patio, Reynolds A Bolt On w/ Ext. Covered Patio, Reynolds A Bolt On w/ Morning Room, and Reynolds A Bolt On w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds B Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_B2_BoltOn.width
        sheet.cell(row=4,column=count).value = Reynolds_B2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_B2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Reynolds_B2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Reynolds_B2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_B2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_B2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds B Bolt On w/ Patio and Reynolds B Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds B Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds B Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds B Bolt On w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_B2_BoltOn.width
        sheet.cell(row=4,column=count).value = Reynolds_B2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_B2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Reynolds_B2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Reynolds_B2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_B2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_B2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds B Bolt On w/ Covered Patio, Reynolds B Bolt On w/ Ext. Covered Patio, Reynolds B Bolt On w/ Morning Room, and Reynolds B Bolt On w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds C Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_C2_BoltOn.width
        sheet.cell(row=4,column=count).value = Reynolds_C2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_C2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Reynolds_C2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Reynolds_C2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_C2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_C2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds C Bolt On w/ Patio and Reynolds C Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds C Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds C Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds C Bolt On w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_C2_BoltOn.width
        sheet.cell(row=4,column=count).value = Reynolds_C2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_C2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Reynolds_C2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Reynolds_C2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_C2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_C2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds C Bolt On w/ Covered Patio, Reynolds C Bolt On w/ Ext. Covered Patio, Reynolds C Bolt On w/ Morning Room, and Reynolds C Bolt On w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds D Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_D2_BoltOn.width
        sheet.cell(row=4,column=count).value = Reynolds_D2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_D2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Reynolds_D2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Reynolds_D2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_D2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_D2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds D Bolt On w/ Patio and Reynolds D Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Reynolds D Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Reynolds D Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Reynolds D Bolt On w/ Ext. Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Reynolds_D2_BoltOn.width
        sheet.cell(row=4,column=count).value = Reynolds_D2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Reynolds_D2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Reynolds_D2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Reynolds_D2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Reynolds_D2_BoltOn.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Reynolds_D2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Reynolds D Bolt On w/ Covered Patio, Reynolds D Bolt On w/ Ext. Covered Patio, Reynolds D Bolt On w/ Morning Room, and Reynolds D Bolt On w/ Ext. Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Sienna A w/ Porch"):
        sheet.cell(row=3,column=count).value = Sienna_A.width
        sheet.cell(row=4,column=count).value = Sienna_A.length
        for j in range(8, sheet.max_row+1):
                           
            if Sienna_A.width <= sheet2.cell(row=j,column=2).value and Sienna_A.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Sienna_A.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Sienna A w/ Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Sienna A w/ Morning Room Patio/ Covered Patio/ Screened Patio") or (sheet.cell(row=7,column=count).value == "Sienna A w/ Ext. Morning Room Patio") or (sheet.cell(row=7,column=count).value == "Sienna A w/ Screened Porch"):
        sheet.cell(row=3,column=count).value = Sienna_A.width
        sheet.cell(row=4,column=count).value = Sienna_A.cp
        for j in range(8, sheet.max_row+1):
                           
            if Sienna_A.width <= sheet2.cell(row=j,column=2).value and Sienna_A.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Sienna_A.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Sienna A w/ Morning Room Patio/ Covered Patio/ Screened Patio, Sienna A w/ Ext. Morning Room Patio, and Sienna A w/ Screened Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Sienna B & C w/ Porch"):
        sheet.cell(row=3,column=count).value = Sienna_B_C.width
        sheet.cell(row=4,column=count).value = Sienna_B_C.length
        for j in range(8, sheet.max_row+1):
                           
            if Sienna_B_C.width <= sheet2.cell(row=j,column=2).value and Sienna_B_C.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Sienna_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Sienna B & C w/ Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Sienna B & C w/ Morning Room Patio/ Covered Patio/ Screened Patio") or (sheet.cell(row=7,column=count).value == "Sienna B & C w/ Ext. Morning Room Patio") or (sheet.cell(row=7,column=count).value == "Sienna B & C w/ Screened Porch"):
        sheet.cell(row=3,column=count).value = Sienna_B_C.width
        sheet.cell(row=4,column=count).value = Sienna_B_C.cp
        for j in range(8, sheet.max_row+1):
                           
            if Sienna_B_C.width <= sheet2.cell(row=j,column=2).value and Sienna_B_C.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Sienna_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Sienna B & C w/ Morning Room Patio/ Covered Patio/ Screened Patio, Sienna B & C w/ Ext. Morning Room Patio, and Sienna B & C w/ Screened Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Sinclair A, B, & C w/ Patio"):
        sheet.cell(row=3,column=count).value = Sinclair_A1_A3_B1_B3_C1_C3.width
        sheet.cell(row=4,column=count).value = Sinclair_A1_A3_B1_B3_C1_C3.length
        for j in range(8, sheet.max_row+1):
                           
            if Sinclair_A1_A3_B1_B3_C1_C3.width <= sheet2.cell(row=j,column=2).value and Sinclair_A1_A3_B1_B3_C1_C3.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Sinclair_A1_A3_B1_B3_C1_C3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Sinclair A, B, & C w/ Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Sinclair A, B, & C w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Sinclair_A1_A3_B1_B3_C1_C3.width
        sheet.cell(row=4,column=count).value = Sinclair_A1_A3_B1_B3_C1_C3.cp
        for j in range(8, sheet.max_row+1):
                           
            if Sinclair_A1_A3_B1_B3_C1_C3.width <= sheet2.cell(row=j,column=2).value and Sinclair_A1_A3_B1_B3_C1_C3.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Sinclair_A1_A3_B1_B3_C1_C3.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Sinclair A, B, & C w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Tea Olive A & B w/ Patio") or (sheet.cell(row=7,column=count).value == "Tea Olive A & B w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Tea_Olive_A_B.width
        sheet.cell(row=4,column=count).value = Tea_Olive_A_B.length
        for j in range(8, sheet.max_row+1):
                           
            if Tea_Olive_A_B.width <= sheet2.cell(row=j,column=2).value and Tea_Olive_A_B.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Tea_Olive_A_B.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Tea Olive A & B w/ Patio and Tea Olive A & B w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
                
    elif (sheet.cell(row=7,column=count).value == "Tea Olive C w/ Patio") or (sheet.cell(row=7,column=count).value == "Tea Olive C w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Tea_Olive_C.width
        sheet.cell(row=4,column=count).value = Tea_Olive_C.length
        for j in range(8, sheet.max_row+1):
                           
            if Tea_Olive_C.width <= sheet2.cell(row=j,column=2).value and Tea_Olive_C.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Tea_Olive_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Tea Olive C w/ Patio and Tea Olive C w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Tea Olive A & B Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Tea Olive A & B Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Tea_Olive_A_B_Side.width
        sheet.cell(row=4,column=count).value = Tea_Olive_A_B_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Tea_Olive_A_B_Side.width <= sheet2.cell(row=j,column=2).value and Tea_Olive_A_B_Side.cp <= sheet2.cell(row=j,column=3).value and Tea_Olive_A_B_Side.side <= (sheet2.cell(row=j,column=5).value + ((Tea_Olive_A_B_Side.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Tea_Olive_A_B_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Tea Olive A & B Side Entry w/ Patio and Tea Olive A & B Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Tea Olive C Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Tea Olive C Side Entry w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Tea_Olive_C_Side.width
        sheet.cell(row=4,column=count).value = Tea_Olive_C_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Tea_Olive_C_Side.width <= sheet2.cell(row=j,column=2).value and Tea_Olive_C_Side.cp <= sheet2.cell(row=j,column=3).value and Tea_Olive_C_Side.side <= (sheet2.cell(row=j,column=5).value + ((Tea_Olive_C_Side.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Tea_Olive_C_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Tea Olive C Side Entry w/ Patio and Tea Olive C Side Entry w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Tea Olive A & B Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Tea Olive A & B Bolt On w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Tea_Olive_A_B_Side.width
        sheet.cell(row=4,column=count).value = Tea_Olive_A_B_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Tea_Olive_A_B_Side.width <= sheet2.cell(row=j,column=2).value and Tea_Olive_A_B_Side.cp <= sheet2.cell(row=j,column=3).value and Tea_Olive_A_B_Side.side <= (sheet2.cell(row=j,column=5).value + ((Tea_Olive_A_B_Side.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Tea_Olive_A_B_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Tea Olive A & B Bolt On w/ Patio and Tea Olive A & B Bolt On w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Tea Olive C Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Tea Olive C Bolt On w/ Covered Patio"):
        sheet.cell(row=3,column=count).value = Tea_Olive_C_Side.width
        sheet.cell(row=4,column=count).value = Tea_Olive_C_Side.cp
        for j in range(8, sheet.max_row+1):
                           
            if Tea_Olive_C_Side.width <= sheet2.cell(row=j,column=2).value and Tea_Olive_C_Side.cp <= sheet2.cell(row=j,column=3).value and Tea_Olive_C_Side.side <= (sheet2.cell(row=j,column=5).value + ((Tea_Olive_C_Side.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Tea_Olive_C_Side.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Tea Olive C Bolt On w/ Patio and Tea Olive C Bolt On w/ Covered Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow A w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow A w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_A2.width
        sheet.cell(row=4,column=count).value = Willow_A2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_A2.width <= sheet2.cell(row=j,column=2).value and Willow_A2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_A2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow A w/ Patio and Willow A w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow A w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow A w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow A w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow A w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_A2.width
        sheet.cell(row=4,column=count).value = Willow_A2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_A2.width <= sheet2.cell(row=j,column=2).value and Willow_A2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_A2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow A w/ Covered Patio, Willow A w/ Ext. Covered Patio, Willow A w/ Morning Room, and Willow A w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow B, C, & D w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_B2_C2_D2.width
        sheet.cell(row=4,column=count).value = Willow_B2_C2_D2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_B2_C2_D2.width <= sheet2.cell(row=j,column=2).value and Willow_B2_C2_D2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_B2_C2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow B, C, & D w/ Patio and Willow B, C, & D w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow B, C, & D w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_B2_C2_D2.width
        sheet.cell(row=4,column=count).value = Willow_B2_C2_D2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_B2_C2_D2.width <= sheet2.cell(row=j,column=2).value and Willow_B2_C2_D2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_B2_C2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow B, C, & D w/ Covered Patio, Willow B, C, & D w/ Ext. Covered Patio, Willow B, C, & D w/ Morning Room, and Willow B, C, & D w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow E w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow E w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_E2.width
        sheet.cell(row=4,column=count).value = Willow_E2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_E2.width <= sheet2.cell(row=j,column=2).value and Willow_E2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_E2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow E w/ Patio and Willow E w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow E w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow E w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow E w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow E w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_E2.width
        sheet.cell(row=4,column=count).value = Willow_E2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_E2.width <= sheet2.cell(row=j,column=2).value and Willow_E2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_E2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow E w/ Covered Patio, Willow E w/ Ext. Covered Patio, Willow E w/ Morning Room, and Willow E w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow F w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow F w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_F2.width
        sheet.cell(row=4,column=count).value = Willow_F2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_F2.width <= sheet2.cell(row=j,column=2).value and Willow_F2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_F2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow F w/ Patio and Willow F w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow F w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow F w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow F w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow F w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_F2.width
        sheet.cell(row=4,column=count).value = Willow_F2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_F2.width <= sheet2.cell(row=j,column=2).value and Willow_F2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_F2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow F w/ Covered Patio, Willow F w/ Ext. Covered Patio, Willow F w/ Morning Room, and Willow F w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow G w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow G w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_G2.width
        sheet.cell(row=4,column=count).value = Willow_G2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_G2.width <= sheet2.cell(row=j,column=2).value and Willow_G2.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_G2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow G w/ Patio and Willow G w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow G w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow G w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow G w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow G w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_G2.width
        sheet.cell(row=4,column=count).value = Willow_G2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_G2.width <= sheet2.cell(row=j,column=2).value and Willow_G2.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_G2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow G w/ Covered Patio, Willow G w/ Ext. Covered Patio, Willow G w/ Morning Room, and Willow G w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow A Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow A Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_A2.width
        sheet.cell(row=4,column=count).value = Willow_A2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_A2.width <= sheet2.cell(row=j,column=2).value and Willow_A2.length <= sheet2.cell(row=j,column=3).value and Willow_A2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_A2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_A2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow A Side Entry w/ Patio and Willow A Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow A Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow A Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow A Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow A Side Entry w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_A2.width
        sheet.cell(row=4,column=count).value = Willow_A2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_A2.width <= sheet2.cell(row=j,column=2).value and Willow_A2.cp <= sheet2.cell(row=j,column=3).value and Willow_A2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_A2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_A2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow A Side Entry w/ Covered Patio, Willow A Side Entry w/ Ext. Covered Patio, Willow A Side Entry w/ Morning Room, AND Willow A Side Entry w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow B, C, & D Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_B2_C2_D2.width
        sheet.cell(row=4,column=count).value = Willow_B2_C2_D2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_B2_C2_D2.width <= sheet2.cell(row=j,column=2).value and Willow_B2_C2_D2.length <= sheet2.cell(row=j,column=3).value and Willow_B2_C2_D2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_B2_C2_D2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_B2_C2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow B, C, & D Side Entry w/ Patio and Willow B, C, & D Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow B, C, & D Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D Side Entry w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_B2_C2_D2.width
        sheet.cell(row=4,column=count).value = Willow_B2_C2_D2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_B2_C2_D2.width <= sheet2.cell(row=j,column=2).value and Willow_B2_C2_D2.cp <= sheet2.cell(row=j,column=3).value and Willow_B2_C2_D2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_B2_C2_D2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_B2_C2_D2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow B, C, & D Side Entry w/ Covered Patio, Willow B, C, & D Side Entry w/ Ext. Covered Patio, Willow B, C, & D Side Entry w/ Morning Room, and Willow B, C, & D Side Entry w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow E Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow E Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_E2.width
        sheet.cell(row=4,column=count).value = Willow_E2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_E2.width <= sheet2.cell(row=j,column=2).value and Willow_E2.length <= sheet2.cell(row=j,column=3).value and Willow_E2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_E2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_E2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow E Side Entry w/ Patio and Willow E Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow E Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow E Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow E Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow E Side Entry w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_E2.width
        sheet.cell(row=4,column=count).value = Willow_E2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_E2.width <= sheet2.cell(row=j,column=2).value and Willow_E2.cp <= sheet2.cell(row=j,column=3).value and Willow_E2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_E2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_E2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow E Side Entry w/ Covered Patio, Willow E Side Entry w/ Ext. Covered Patio, Willow E Side Entry w/ Morning Room, and Willow E Side Entry w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow F Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow F Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_E2.width
        sheet.cell(row=4,column=count).value = Willow_E2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_F2.width <= sheet2.cell(row=j,column=2).value and Willow_F2.length <= sheet2.cell(row=j,column=3).value and Willow_F2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_F2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_F2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow F Side Entry w/ Patio and Willow F Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow F Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow F Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow F Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow F Side Entry w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_F2.width
        sheet.cell(row=4,column=count).value = Willow_F2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_F2.width <= sheet2.cell(row=j,column=2).value and Willow_F2.cp <= sheet2.cell(row=j,column=3).value and Willow_F2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_F2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_F2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow F Side Entry w/ Covered Patio, Willow F Side Entry w/ Ext. Covered Patio, Willow F Side Entry w/ Morning Room, and Willow F Side Entry w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow G Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow G Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_G2.width
        sheet.cell(row=4,column=count).value = Willow_G2.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_G2.width <= sheet2.cell(row=j,column=2).value and Willow_G2.length <= sheet2.cell(row=j,column=3).value and Willow_G2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_G2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_G2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow G Side Entry w/ Patio and Willow G Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow G Side Entry w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow G Side Entry w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow G Side Entry w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow G Side Entry w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_G2.width
        sheet.cell(row=4,column=count).value = Willow_G2.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_G2.width <= sheet2.cell(row=j,column=2).value and Willow_G2.cp <= sheet2.cell(row=j,column=3).value and Willow_G2.side <= (sheet2.cell(row=j,column=5).value + ((Willow_G2.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_G2.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow G Side Entry w/ Covered Patio, Willow G Side Entry w/ Ext. Covered Patio, Willow G Side Entry w/ Morning Room, and Willow G Side Entry w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow A Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow A Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_A2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_A2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_A2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_A2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Willow_A2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_A2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_A2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow A Bolt On w/ Patio and Willow A Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow A Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow A Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow A Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow A Bolt On w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_A2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_A2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_A2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_A2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Willow_A2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_A2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_A2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow A Bolt On w/ Covered Patio, Willow A Bolt On w/ Ext. Covered Patio, Willow A Bolt On w/ Morning Room, and Willow A Bolt On w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow B, C, & D Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_B2_C2_D2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_B2_C2_D2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_B2_C2_D2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_B2_C2_D2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Willow_B2_C2_D2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_B2_C2_D2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_B2_C2_D2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow B, C, & D Bolt On w/ Patio and Willow B, C, & D Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow B, C, & D Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow B, C, & D Bolt On w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_B2_C2_D2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_B2_C2_D2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_B2_C2_D2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_B2_C2_D2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Willow_B2_C2_D2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_B2_C2_D2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_B2_C2_D2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow B, C, & D Bolt On w/ Covered Patio, Willow B, C, & D Bolt On w/ Ext. Covered Patio, Willow B, C, & D Bolt On w/ Morning Room, and Willow B, C, & D Bolt On w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow E Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow E Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_E2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_E2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_E2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_E2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Willow_E2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_E2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_E2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow E Bolt On w/ Patio and Willow E Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow E Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow E Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow E Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow E Bolt On w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_E2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_E2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_E2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_E2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Willow_E2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_E2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_E2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow E Bolt On w/ Covered Patio, Willow E Bolt On w/ Ext. Covered Patio, Willow E Bolt On w/ Morning Room, and Willow E Bolt On w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow F Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow F Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_F2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_F2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_F2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_F2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Willow_F2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_F2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_F2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow F Bolt On w/ Patio and Willow F Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow F Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow F Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow F Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow F Bolt On w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_E2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_E2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_F2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_F2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Willow_F2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_F2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_F2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow F Bolt On w/ Covered Patio, Willow F Bolt On w/ Ext. Covered Patio, Willow F Bolt On w/ Morning Room, and Willow F Bolt On w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow G Bolt On w/ Patio") or (sheet.cell(row=7,column=count).value == "Willow G Bolt On w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Willow_G2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_G2_BoltOn.length
        for j in range(8, sheet.max_row+1):
                           
            if Willow_G2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_G2_BoltOn.length <= sheet2.cell(row=j,column=3).value and Willow_G2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_G2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_G2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow G Bolt On w/ Patio and Willow G Bolt On w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Willow G Bolt On w/ Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow G Bolt On w/ Ext. Covered Patio") or (sheet.cell(row=7,column=count).value == "Willow G Bolt On w/ Morning Room") or (sheet.cell(row=7,column=count).value == "Willow G Bolt On w/ Morning Room Patio"):
        sheet.cell(row=3,column=count).value = Willow_G2_BoltOn.width
        sheet.cell(row=4,column=count).value = Willow_G2_BoltOn.cp
        for j in range(8, sheet.max_row+1):
                           
            if Willow_G2_BoltOn.width <= sheet2.cell(row=j,column=2).value and Willow_G2_BoltOn.cp <= sheet2.cell(row=j,column=3).value and Willow_G2_BoltOn.side <= (sheet2.cell(row=j,column=5).value + ((Willow_G2_BoltOn.width-sheet2.cell(row=j,column=2).value))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Willow_G2_BoltOn.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Willow G Bolt On w/ Covered Patio, Willow G Bolt On w/ Ext. Covered Patio, Willow G Bolt On w/ Morning Room, and Willow G Bolt On w/ Morning Room Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Winston A, B, & C w/ Patio") or (sheet.cell(row=7,column=count).value == "Winston A, B, & C w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Winston_A_B_C.width
        sheet.cell(row=4,column=count).value = Winston_A_B_C.length
        for j in range(8, sheet.max_row+1):
                           
            if Winston_A_B_C.width <= sheet2.cell(row=j,column=2).value and Winston_A_B_C.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Winston_A_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Winston A, B, & C w/ Patio and Winston A, B, & C w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Winston A, B, & C w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Winston_A_B_C.width
        sheet.cell(row=4,column=count).value = Winston_A_B_C.cp
        for j in range(8, sheet.max_row+1):
                           
            if Winston_A_B_C.width <= sheet2.cell(row=j,column=2).value and Winston_A_B_C.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Winston_A_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Winston A, B, & C w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Winston A, B, & C Side Entry w/ Patio") or (sheet.cell(row=7,column=count).value == "Winston A, B, & C Side Entry w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Winston_A_B_C.width
        sheet.cell(row=4,column=count).value = Winston_A_B_C.cp
        for j in range(8, sheet.max_row+1):
                           
            if Winston_A_B_C.width <= sheet2.cell(row=j,column=2).value and Winston_A_B_C.cp <= sheet2.cell(row=j,column=3).value and Winston_A_B_C.side <= (sheet2.cell(row=j,column=5).value + ((sheet2.cell(row=j,column=2).value-Winston_A_B_C.width))):
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Winston_A_B_C.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Winston A, B, & C Side Entry w/ Patio and Winston A, B, & C Side Entry w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Winston A, B, & C 3 Front w/ Patio") or (sheet.cell(row=7,column=count).value == "Winston A, B, & C 3 Front w/ Ext. Patio"):
        sheet.cell(row=3,column=count).value = Winston_A_B_C_3_Front.width
        sheet.cell(row=4,column=count).value = Winston_A_B_C_3_Front.length
        for j in range(8, sheet.max_row+1):
                           
            if Winston_A_B_C_3_Front.width <= sheet2.cell(row=j,column=2).value and Winston_A_B_C_3_Front.length <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Winston_A_B_C_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Winston A, B, & C 3 Front w/ Patio and Winston A, B, & C 3 Front w/ Ext. Patio does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    elif (sheet.cell(row=7,column=count).value == "Winston A, B, & C 3 Front w/ Covered Porch"):
        sheet.cell(row=3,column=count).value = Winston_A_B_C_3_Front.width
        sheet.cell(row=4,column=count).value = Winston_A_B_C_3_Front.cp
        for j in range(8, sheet.max_row+1):
                           
            if Winston_A_B_C_3_Front.width <= sheet2.cell(row=j,column=2).value and Winston_A_B_C_3_Front.cp <= sheet2.cell(row=j,column=3).value:
                    
                sheet.cell(row=j,column=count).value = sheet2.cell(row=j,column=4).value + (sheet2.cell(row=j,column=3).value - Winston_A_B_C_3_Front.length)
                sheet.cell(row=j, column=count).fill = greenFill
                sheet.cell(row=j, column=count).border = thin_border
            else:
                sheet.cell(row=j, column=count).fill = redFill
                sheet.cell(row=j, column=count).border = thin_border
                print(f"Winston A, B, & C 3 Front w/ Covered Porch does not fit in lot {sheet.cell(row=j, column=1).value}")
                
    
                
    
                
    
                
wb.save('1_Lot Fit Matrix with Exposed Columns.xlsx')
                